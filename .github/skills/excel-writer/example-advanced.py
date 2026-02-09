#!/usr/bin/env python3
"""
Advanced Excel generation example with multiple sheets and charts
Run with: uvx --with openpyxl python example-advanced.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime, timedelta

def create_quarterly_report(filename: str = "quarterly_report.xlsx"):
    """Create a multi-sheet quarterly report with charts"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    summary_ws = wb.create_sheet("Summary")
    monthly_ws = wb.create_sheet("Monthly Details")
    products_ws = wb.create_sheet("Products")
    
    # ===== SUMMARY SHEET =====
    summary_ws['A1'] = "Q1 2025 Sales Summary"
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws.merge_cells('A1:D1')
    
    summary_headers = ["Month", "Sales", "Costs", "Profit"]
    summary_ws.append([])  # Empty row
    summary_ws.append(summary_headers)
    
    # Style summary headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in summary_ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Summary data
    summary_data = [
        ["January", 125000, 85000, "=B4-C4"],
        ["February", 145000, 95000, "=B5-C5"],
        ["March", 165000, 105000, "=B6-C6"],
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Add totals
    summary_ws.append(["TOTAL", "=SUM(B4:B6)", "=SUM(C4:C6)", "=SUM(D4:D6)"])
    summary_ws['A7'].font = Font(bold=True)
    summary_ws['B7'].font = Font(bold=True)
    summary_ws['C7'].font = Font(bold=True)
    summary_ws['D7'].font = Font(bold=True)
    
    # Format currency
    for row in range(4, 8):
        for col in ['B', 'C', 'D']:
            summary_ws[f'{col}{row}'].number_format = '$#,##0'
    
    # Add bar chart
    chart = BarChart()
    chart.title = "Monthly Profit"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Profit ($)"
    
    data = Reference(summary_ws, min_col=4, min_row=3, max_row=6)
    categories = Reference(summary_ws, min_col=1, min_row=4, max_row=6)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    summary_ws.add_chart(chart, "F3")
    
    # ===== MONTHLY DETAILS SHEET =====
    monthly_ws['A1'] = "Monthly Sales Details"
    monthly_ws['A1'].font = Font(bold=True, size=14)
    monthly_ws.merge_cells('A1:E1')
    
    detail_headers = ["Date", "Product", "Quantity", "Unit Price", "Total"]
    monthly_ws.append([])
    monthly_ws.append(detail_headers)
    
    # Style headers
    for cell in monthly_ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Generate sample daily data
    base_date = datetime(2025, 1, 1)
    products = ["Laptop", "Desktop", "Tablet", "Phone", "Monitor"]
    
    for i in range(30):
        date = base_date + timedelta(days=i)
        product = products[i % len(products)]
        quantity = (i % 10) + 1
        price = 500 + (i * 50)
        monthly_ws.append([date, product, quantity, price, f"=C{i+4}*D{i+4}"])
    
    # Format columns
    monthly_ws.column_dimensions['A'].width = 12
    monthly_ws.column_dimensions['B'].width = 15
    monthly_ws.column_dimensions['C'].width = 10
    monthly_ws.column_dimensions['D'].width = 12
    monthly_ws.column_dimensions['E'].width = 12
    
    # Format data
    for row in range(4, monthly_ws.max_row + 1):
        monthly_ws[f'A{row}'].number_format = 'YYYY-MM-DD'
        monthly_ws[f'D{row}'].number_format = '$#,##0.00'
        monthly_ws[f'E{row}'].number_format = '$#,##0.00'
    
    # ===== PRODUCTS SHEET =====
    products_ws['A1'] = "Product Performance"
    products_ws['A1'].font = Font(bold=True, size=14)
    products_ws.merge_cells('A1:C1')
    
    product_headers = ["Product", "Units Sold", "Revenue"]
    products_ws.append([])
    products_ws.append(product_headers)
    
    # Style headers
    for cell in products_ws[3]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Product summary data
    product_data = [
        ["Laptop", 45, 44995],
        ["Desktop", 32, 38400],
        ["Tablet", 78, 23400],
        ["Phone", 92, 36800],
        ["Monitor", 55, 16500],
    ]
    
    for row in product_data:
        products_ws.append(row)
    
    # Format currency
    for row in range(4, products_ws.max_row + 1):
        products_ws[f'C{row}'].number_format = '$#,##0'
    
    # Add pie chart for product revenue
    pie = PieChart()
    pie.title = "Revenue by Product"
    
    data = Reference(products_ws, min_col=3, min_row=3, max_row=products_ws.max_row)
    labels = Reference(products_ws, min_col=1, min_row=4, max_row=products_ws.max_row)
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    products_ws.add_chart(pie, "E3")
    
    # Set column widths
    products_ws.column_dimensions['A'].width = 15
    products_ws.column_dimensions['B'].width = 12
    products_ws.column_dimensions['C'].width = 12
    
    # Save workbook
    wb.save(filename)
    print(f"✓ Advanced Excel report created: {filename}")
    print(f"  - Summary sheet with bar chart")
    print(f"  - Monthly details with 30 transactions")
    print(f"  - Product performance with pie chart")

if __name__ == "__main__":
    create_quarterly_report()
