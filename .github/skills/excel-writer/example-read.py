#!/usr/bin/env python3
"""
Excel reading example - reads data from an Excel file and displays/exports it
Run with: uvx --with openpyxl python example-read.py
"""

from openpyxl import load_workbook
from pathlib import Path
import json
import sys

def read_and_display_excel(filename: str):
    """Read an Excel file and display its contents"""
    
    if not Path(filename).exists():
        print(f"❌ File not found: {filename}")
        print(f"   Please make sure the file exists.")
        return None
    
    try:
        # Load the workbook (data_only=True gets calculated formula values)
        wb = load_workbook(filename, data_only=True)
        
        print(f"\n{'='*60}")
        print(f"📊 Excel File: {filename}")
        print(f"{'='*60}\n")
        print(f"Available sheets: {', '.join(wb.sheetnames)}\n")
        
        all_data = {}
        
        # Read each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"\n📄 Sheet: '{sheet_name}'")
            print(f"   Dimensions: {ws.dimensions}")
            print(f"   Total rows: {ws.max_row}")
            print(f"   Total columns: {ws.max_column}")
            
            # Read all rows
            sheet_data = []
            for row in ws.iter_rows(values_only=True):
                sheet_data.append(row)
            
            all_data[sheet_name] = sheet_data
            
            # Display sample data
            if sheet_data:
                print(f"\n   Sample data (first {min(5, len(sheet_data))} rows):")
                print(f"   {'-'*50}")
                for i, row in enumerate(sheet_data[:5], 1):
                    # Format row nicely
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    print(f"   {i:2d}. {row_str}")
                
                if len(sheet_data) > 5:
                    print(f"   ... and {len(sheet_data) - 5} more rows")
            else:
                print("   (empty sheet)")
        
        print(f"\n{'='*60}\n")
        return all_data
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

def excel_to_dict_list(filename: str, sheet_name: str = None):
    """Convert Excel to list of dictionaries (using first row as headers)"""
    
    wb = load_workbook(filename, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    
    # Get all rows
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        return []
    
    # First row as headers
    headers = rows[0]
    
    # Convert remaining rows to dictionaries
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
        data.append(row_dict)
    
    return data

def excel_to_json(filename: str, output_file: str = "output.json", sheet_name: str = None):
    """Convert Excel file to JSON"""
    
    print(f"\n🔄 Converting Excel to JSON...")
    
    data = excel_to_dict_list(filename, sheet_name)
    
    # Convert to JSON (handle dates and None values)
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    
    print(f"✓ Successfully converted to JSON: {output_file}")
    print(f"  Total records: {len(data)}")
    
    # Show sample
    if data:
        print(f"\n  Sample record:")
        print(f"  {json.dumps(data[0], indent=4, default=str)}")

def read_specific_range(filename: str, sheet_name: str, range_str: str):
    """Read a specific cell range from Excel"""
    
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    
    print(f"\n📋 Reading range '{range_str}' from sheet '{sheet_name}':")
    
    data = []
    for row in ws[range_str]:
        row_data = [cell.value for cell in row]
        data.append(row_data)
        print(f"   {row_data}")
    
    return data

if __name__ == "__main__":
    # Example 1: Read the quarterly report we created earlier
    example_file = "quarterly_report.xlsx"
    
    if Path(example_file).exists():
        print("Example 1: Reading entire Excel file")
        data = read_and_display_excel(example_file)
        
        print("\n" + "="*60)
        print("Example 2: Converting to JSON")
        excel_to_json(example_file, "quarterly_data.json", "Summary")
        
        print("\n" + "="*60)
        print("Example 3: Reading specific range")
        read_specific_range(example_file, "Summary", "A3:D7")
        
    else:
        print(f"ℹ️  File '{example_file}' not found.")
        print(f"   Run the advanced example first to create it:")
        print(f"   uvx --with openpyxl python example-advanced.py")
        print()
        print(f"   Or create it and try reading any Excel file:")
        print(f"   Just modify the 'example_file' variable in this script.")
