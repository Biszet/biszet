#!/usr/bin/env python3
"""
Excel modification example - reads an existing Excel file and adds new content
Run with: uvx --with openpyxl python example-modify.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from pathlib import Path
from datetime import datetime

def modify_excel_file(input_file: str, output_file: str = None):
    """
    Read an existing Excel file and add a new analysis sheet
    """
    
    if not Path(input_file).exists():
        print(f"❌ File not found: {input_file}")
        print(f"   Please run example-advanced.py first to create it.")
        return
    
    # Use input file name if no output specified
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_modified.xlsx')
    
    print(f"\n📂 Loading Excel file: {input_file}")
    
    # Load existing workbook
    wb = load_workbook(input_file)
    
    print(f"   Existing sheets: {', '.join(wb.sheetnames)}")
    
    # Create new analysis sheet
    if "Analysis" in wb.sheetnames:
        del wb["Analysis"]  # Remove if exists
    
    analysis_ws = wb.create_sheet("Analysis")
    
    print(f"   ✓ Created new sheet: 'Analysis'")
    
    # Add title
    analysis_ws['A1'] = "Data Analysis Report"
    analysis_ws['A1'].font = Font(bold=True, size=16)
    analysis_ws.merge_cells('A1:D1')
    
    analysis_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    analysis_ws['A2'].font = Font(italic=True, size=10)
    
    # Add summary statistics
    analysis_ws.append([])  # Empty row
    analysis_ws.append(["Metric", "Q1 Total", "Average", "Best Month"])
    
    # Style header
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    for cell in analysis_ws[4]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Add calculated metrics (referencing Summary sheet)
    metrics = [
        ["Sales", "=Summary!B7", "=Summary!B7/3", "=MAX(Summary!B4:B6)"],
        ["Costs", "=Summary!C7", "=Summary!C7/3", "=MAX(Summary!C4:C6)"],
        ["Profit", "=Summary!D7", "=Summary!D7/3", "=MAX(Summary!D4:D6)"],
    ]
    
    for row in metrics:
        analysis_ws.append(row)
    
    # Format currency columns
    for row in range(5, 8):
        for col in ['B', 'C', 'D']:
            analysis_ws[f'{col}{row}'].number_format = '$#,##0'
    
    # Set column widths
    analysis_ws.column_dimensions['A'].width = 15
    analysis_ws.column_dimensions['B'].width = 12
    analysis_ws.column_dimensions['C'].width = 12
    analysis_ws.column_dimensions['D'].width = 12
    
    # Add trend analysis section
    analysis_ws.append([])  # Empty row
    analysis_ws.append(["Month", "Growth Rate"])
    
    # Header for trend section
    for cell in analysis_ws[9]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate growth rates
    trend_data = [
        ["Jan → Feb", "=(Summary!B5-Summary!B4)/Summary!B4"],
        ["Feb → Mar", "=(Summary!B6-Summary!B5)/Summary!B5"],
    ]
    
    for row in trend_data:
        analysis_ws.append(row)
    
    # Format percentages
    for row in range(10, 12):
        analysis_ws[f'B{row}'].number_format = '0.0%'
    
    # Add a trend chart
    chart = LineChart()
    chart.title = "Sales Trend"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Sales ($)"
    
    # Reference data from Summary sheet
    data = Reference(wb["Summary"], min_col=2, min_row=3, max_row=6)
    categories = Reference(wb["Summary"], min_col=1, min_row=4, max_row=6)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    analysis_ws.add_chart(chart, "F4")
    
    # Add insights section
    analysis_ws['A14'] = "Key Insights:"
    analysis_ws['A14'].font = Font(bold=True, size=12)
    
    insights = [
        "• Revenue shows consistent growth across Q1",
        "• Average monthly profit exceeds $50,000",
        "• March represents the strongest performing month",
        "• Cost efficiency improved throughout the quarter",
    ]
    
    for i, insight in enumerate(insights, start=15):
        analysis_ws[f'A{i}'] = insight
        analysis_ws[f'A{i}'].alignment = Alignment(wrap_text=True)
    
    # Save modified workbook
    wb.save(output_file)
    
    print(f"\n✓ Excel file modified successfully!")
    print(f"   Output file: {output_file}")
    print(f"   New sheets: {', '.join(wb.sheetnames)}")
    print(f"\n   Added:")
    print(f"   • Analysis sheet with calculated metrics")
    print(f"   • Cross-sheet formulas referencing Summary data")
    print(f"   • Growth rate calculations")
    print(f"   • Trend chart")
    print(f"   • Key insights section")

if __name__ == "__main__":
    # Modify the quarterly report
    input_file = "quarterly_report.xlsx"
    
    if Path(input_file).exists():
        modify_excel_file(input_file)
    else:
        print(f"⚠️  Input file '{input_file}' not found.")
        print(f"   Please create it first by running:")
        print(f"   uvx --with openpyxl python example-advanced.py")
