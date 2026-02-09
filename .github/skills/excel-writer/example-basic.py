#!/usr/bin/env python3
"""
Basic Excel generation example
Run with: uvx --with openpyxl python example-basic.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_sales_report(filename: str = "sales_report.xlsx"):
    """Create a simple sales report Excel file"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers
    headers = ["Date", "Product", "Quantity", "Unit Price", "Total"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data
    sales_data = [
        [datetime(2025, 1, 15), "Laptop", 5, 999.99, "=C2*D2"],
        [datetime(2025, 1, 16), "Mouse", 15, 29.99, "=C3*D3"],
        [datetime(2025, 1, 17), "Keyboard", 10, 79.99, "=C4*D4"],
        [datetime(2025, 1, 18), "Monitor", 8, 299.99, "=C5*D5"],
        [datetime(2025, 1, 19), "Webcam", 12, 89.99, "=C6*D6"],
    ]
    
    # Add data
    for row in sales_data:
        ws.append(row)
    
    # Format columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    # Format date column
    for row in range(2, ws.max_row + 1):
        ws[f'A{row}'].number_format = 'YYYY-MM-DD'
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].number_format = '$#,##0.00'
    
    # Add total row
    total_row = ws.max_row + 1
    ws[f'B{total_row}'] = "TOTAL"
    ws[f'B{total_row}'].font = Font(bold=True, size=11)
    ws[f'E{total_row}'] = f'=SUM(E2:E{ws.max_row})'
    ws[f'E{total_row}'].font = Font(bold=True, size=11)
    ws[f'E{total_row}'].number_format = '$#,##0.00'
    
    # Save
    wb.save(filename)
    print(f"✓ Excel file created successfully: {filename}")
    print(f"  - Contains {len(sales_data)} sales records")
    print(f"  - Total calculated automatically")

if __name__ == "__main__":
    create_sales_report()
